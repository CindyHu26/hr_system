import streamlit as st
import pandas as pd
from datetime import datetime
import io
import numpy as np
from openpyxl.styles import PatternFill
from utils_salary_calc import (
    calculate_salary_df,
    save_salary_df,
    get_previous_non_insured_names,
    get_salary_report_for_editing,
    check_salary_records_exist,
    get_item_types,
    save_data_editor_changes,
    batch_update_salary_details_from_excel
)

# --- [CORE FIX] ---
# Modify the function to accept item_types
def dataframe_to_excel(df, item_types, sheet_name='Sheet1', color_map=None):
    """將 DataFrame 轉換為使用 openpyxl 引擎的 Excel 二進位格式，並可選擇性地上色"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        if color_map:
            header_map = {v: k for k, v in color_map['required_cols'].items()}
            
            for col_idx, column_title in enumerate(df.columns, 1):
                original_title = header_map.get(column_title, column_title)
                fill = None
                if item_types.get(original_title) == 'earning': fill = color_map['earning_fill']
                elif item_types.get(original_title) == 'deduction': fill = color_map['deduction_fill']
                elif original_title in ['應發總額', '應扣總額', '實發淨薪']: fill = color_map['total_fill']
                elif '公司負擔' in original_title or '勞退提撥' in original_title: fill = color_map['cost_fill']

                if fill:
                    for row_idx in range(2, len(df) + 2):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill
        
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = length
            
    return output.getvalue()

def generate_payslip_excel(df, item_types):
    """產生用於列印薪資條的Excel報表，並獨立計算合計欄位"""
    payslip_df = df.copy()
    required_cols = {
        '員工姓名': '姓名', '員工編號': '編號', '加保單位': '加保單位', '底薪': '底薪', 
        '加班費': '加班費', '延長工時': '延長工時', '加班費2': '加班費2', '再延長工時': '再延長工時',
        '應發總額': '應付合計', '勞健保': '勞健保', '借支': '借支', '事假': '事假', '病假': '病假', 
        '遲到(分)': '遲到(分)', '遲到': '遲到', '早退(分)': '早退(分)', '早退': '早退',
        '二代健保補充費':'其他', '稅款': '稅款', '應扣總額': '應扣合計', 
        '實發淨薪': '合計', '勞退提撥(公司負擔)': '勞退提撥'
    }
    
    for col in required_cols:
        if col not in payslip_df.columns:
            payslip_df[col] = 0 if '分' not in col and '工時' not in col else 0.0
    
    payslip_earning_cols = [col for col, item_type in item_types.items() if item_type == 'earning' and col in required_cols and col in payslip_df.columns]
    payslip_deduction_cols = [col for col, item_type in item_types.items() if item_type == 'deduction' and col in required_cols and col in payslip_df.columns]
    
    payslip_df['應發總額'] = payslip_df[payslip_earning_cols].sum(axis=1)
    payslip_df['應扣總額'] = payslip_df[payslip_deduction_cols].sum(axis=1)
    payslip_df['實發淨薪'] = payslip_df['應發總額'] + payslip_df['應扣總額']

    final_df = payslip_df[list(required_cols.keys())].rename(columns=required_cols)
    
    color_map = {
        'earning_fill': PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        'deduction_fill': PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        'total_fill': PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
        'cost_fill': PatternFill(start_color="E5E7E9", end_color="E5E7E9", fill_type="solid"),
        'required_cols': required_cols
    }
        
    # [CORE FIX] Pass item_types to the helper function
    return dataframe_to_excel(final_df, item_types, sheet_name='薪資條', color_map=color_map)

def generate_financial_report_excel(df, item_types):
    """產生包含公司成本與總計的完整財務報表"""
    report_df = df.copy()
    numeric_cols = report_df.select_dtypes(include=np.number).columns.tolist()
    total_row = report_df[numeric_cols].sum().to_frame().T
    total_row['員工姓名'] = '*** 總計 ***'
    report_df = pd.concat([report_df, total_row], ignore_index=True)
    
    color_map = {
        'earning_fill': PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        'deduction_fill': PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        'total_fill': PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
        'cost_fill': PatternFill(start_color="E5E7E9", end_color="E5E7E9", fill_type="solid"),
        'required_cols': {col: col for col in df.columns}
    }
            
    # [CORE FIX] Pass item_types to the helper function
    return dataframe_to_excel(report_df, item_types, sheet_name='完整財務報表', color_map=color_map)


def show_page(conn):
    st.header("薪資單產生與管理")

    if 'salary_workflow_step' not in st.session_state: st.session_state.salary_workflow_step = 'initial'
    if 'salary_draft_df' not in st.session_state: st.session_state.salary_draft_df = None

    c1, c2 = st.columns(2)
    today = datetime.now()
    year = c1.number_input("選擇年份", min_value=2020, max_value=today.year + 5, value=today.year)
    month = c2.number_input("選擇月份", min_value=1, max_value=12, value=today.month)

    if 'current_month' not in st.session_state or st.session_state.current_month != (year, month):
        st.session_state.salary_workflow_step = 'initial'
        st.session_state.salary_draft_df = None
        st.session_state.current_month = (year, month)

    if st.session_state.salary_workflow_step in ['initial', 'saved']:
        st.write("---")
        records_exist = check_salary_records_exist(conn, year, month)
        
        if records_exist: st.success(f"✅ {year} 年 {month} 月的薪資單已儲存在資料庫中。")
        else: st.info(f"💡 {year} 年 {month} 月的薪資單尚未產生。")

        if st.button("🚀 產生薪資草稿 (試算)", type="primary"):
            with st.spinner("正在為所有在職員工進行薪資試算..."):
                draft_df, _ = calculate_salary_df(conn, year, month)
                st.session_state.salary_draft_df = draft_df
                st.session_state.salary_workflow_step = 'draft'
                st.rerun()

    elif st.session_state.salary_workflow_step == 'draft':
        st.write("---")
        st.subheader("📝 薪資草稿預覽與確認")
        st.warning("以下為試算結果，尚未存入資料庫。")

        draft_df = st.session_state.salary_draft_df
        if draft_df is not None and not draft_df.empty:
            st.markdown("##### 1. 設定勞健保自理人員")
            all_emp_names_in_draft = draft_df['員工姓名'].tolist()
            previous_non_insured = get_previous_non_insured_names(conn, year, month)
            default_selection = [name for name in previous_non_insured if name in all_emp_names_in_draft]
            
            non_insured_names = st.multiselect("選擇勞健保不計(=0)的員工", options=all_emp_names_in_draft, default=default_selection)

            recalculated_draft, _ = calculate_salary_df(conn, year, month, non_insured_names=non_insured_names)
            st.session_state.salary_draft_df = recalculated_draft
            
            st.markdown("##### 2. 預覽計算結果")
            st.dataframe(recalculated_draft)

            c1_btn, c2_btn, _ = st.columns([1, 1, 3])
            if c1_btn.button("✅ 確認並儲存薪資單", type="primary"):
                with st.spinner("正在將薪資單寫入資料庫..."):
                    save_salary_df(conn, year, month, recalculated_draft)
                    st.session_state.salary_workflow_step = 'saved'
                    st.session_state.salary_draft_df = None
                    st.success("薪資單已成功儲存！")
                    st.rerun()

            if c2_btn.button("❌ 放棄此草稿"):
                st.session_state.salary_workflow_step = 'initial'
                st.session_state.salary_draft_df = None
                st.rerun()

    if st.session_state.salary_workflow_step == 'saved' or (st.session_state.salary_workflow_step == 'initial' and check_salary_records_exist(conn, year, month)):
        display_and_edit_section(conn, year, month)
        display_download_section(conn, year, month)

def display_and_edit_section(conn, year, month):
    st.write("---")
    st.subheader("💵 薪資明細微調 (編輯區)")
    
    with st.expander("🚀 批次上傳津貼/費用 (Excel)"):
        uploaded_file = st.file_uploader("上傳 Excel 檔更新薪資", type="xlsx", key=f"salary_excel_uploader_{year}_{month}")
        if uploaded_file:
            with st.spinner("正在處理上傳的 Excel 檔案..."):
                report = batch_update_salary_details_from_excel(conn, year, month, uploaded_file)
                st.success("批次更新完成！")
                if report["success"]: st.write(f"成功更新 {len(report['success'])} 筆資料。")
                if report["skipped_emp"]: st.warning(f"找不到對應員工，已跳過：{', '.join(report['skipped_emp'])}")
                if report["skipped_item"]: st.warning(f"找不到對應薪資項目，已跳過：{', '.join(report['skipped_item'])}")

    st.markdown("##### 手動編輯薪資項目")
    st.caption("您可以在下表中直接修改數值。修改完成後，請點擊下方的「儲存手動變更」按鈕。")
    
    report_df, _ = get_salary_report_for_editing(conn, year, month)
    
    if not report_df.empty:
        disabled_cols = ["員工姓名", "員工編號", "加保單位", '應發總額', '應扣總額', '實發淨薪', '申報薪資']
        
        edit_key = f'before_edit_df_{year}_{month}'
        if edit_key not in st.session_state:
            st.session_state[edit_key] = report_df.copy()

        edited_df = st.data_editor(report_df, use_container_width=True, disabled=disabled_cols, num_rows="dynamic", key=f"data_editor_{year}_{month}")
        
        has_changes = not edited_df.equals(st.session_state[edit_key])
        
        if has_changes: st.warning("偵測到變更，請點擊儲存。")
        
        if st.button("💾 儲存手動變更", type="primary", disabled=not has_changes):
            with st.spinner("正在儲存您的修改..."):
                try:
                    save_data_editor_changes(conn, year, month, edited_df)
                    st.session_state[edit_key] = edited_df.copy()
                    st.success("手動修改儲存成功！")
                    st.rerun()
                except Exception as e:
                    st.error(f"儲存時發生錯誤: {e}")

def display_download_section(conn, year, month):
    st.write("---")
    st.subheader("📊 報表匯出")
    
    report_df, item_types = get_salary_report_for_editing(conn, year, month)
    if not report_df.empty:
        st.info("您可以下載兩種不同格式的Excel報表，用於薪資條列印或財務總覽。")
        roc_year = year - 1911
        
        c1, c2 = st.columns(2)
        
        with c1:
            payslip_excel_data = generate_payslip_excel(report_df, item_types)
            st.download_button(
                label="📥 下載薪資條報表 (Excel)",
                data=payslip_excel_data,
                file_name=f"薪資條_{roc_year}年{month:02d}月.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with c2:
            financial_excel_data = generate_financial_report_excel(report_df, item_types)
            st.download_button(
                label="📥 下載完整財務報表 (Excel)",
                data=financial_excel_data,
                file_name=f"完整財務報表_{roc_year}年{month:02d}月.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.warning("目前資料庫中沒有可供匯出的薪資紀錄。")