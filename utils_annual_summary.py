# utils_annual_summary.py
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_annual_salary_summary(conn, year: int, item_ids: list):
    """
    根據選定的薪資項目，查詢並彙總指定年度的員工薪資。

    Args:
        conn: 資料庫連線。
        year (int): 要查詢的年份。
        item_ids (list): 要加總的薪資項目ID列表。

    Returns:
        pd.DataFrame: 彙總後的年度薪資報表。
    """
    if not item_ids:
        # 如果沒有選擇任何項目，返回一個帶有正確欄位的空 DataFrame
        return pd.DataFrame(columns=['員工編號', '員工姓名'] + [f'{m}月' for m in range(1, 13)])

    placeholders = ','.join('?' for _ in item_ids)
    query = f"""
    SELECT
        e.hr_code as '員工編號',
        e.name_ch as '員工姓名',
        s.month,
        SUM(sd.amount) as monthly_total
    FROM salary_detail sd
    JOIN salary s ON sd.salary_id = s.id
    JOIN employee e ON s.employee_id = e.id
    WHERE s.year = ? AND sd.salary_item_id IN ({placeholders})
    GROUP BY e.id, s.month
    ORDER BY e.hr_code, s.month;
    """
    params = [year] + item_ids
    df = pd.read_sql_query(query, conn, params=params)

    if df.empty:
        return pd.DataFrame(columns=['員工編號', '員工姓名'] + [f'{m}月' for m in range(1, 13)])

    # 將資料從「長格式」轉換為「寬格式」的樞紐表
    pivot_df = df.pivot_table(
        index=['員工編號', '員工姓名'],
        columns='month',
        values='monthly_total',
        fill_value=0
    ).reset_index()

    # 重新命名欄位，例如 1 -> 1月
    pivot_df.columns = ['員工編號', '員工姓名'] + [f'{col}月' for col in pivot_df.columns[2:]]

    # 確保1到12月的欄位都存在
    for m in range(1, 13):
        month_col = f'{m}月'
        if month_col not in pivot_df.columns:
            pivot_df[month_col] = 0

    # 重新排序欄位
    pivot_df = pivot_df[['員工編號', '員工姓名'] + [f'{m}月' for m in range(1, 13)]]

    return pivot_df

def dataframe_to_styled_excel(df, title, roc_year):
    """
    將 DataFrame 轉換為帶有樣式的 Excel 二進位格式。
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "年度薪資總表"

    # --- 標題 ---
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"民國 {roc_year} 年 {title}"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- 寫入 DataFrame ---
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        ws.append(row)

    # --- 設定樣式 ---
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for col_idx, cell in enumerate(ws[3], 1): # 第3列是標頭
        cell.font = header_font
        cell.fill = header_fill
        # 自動調整欄寬
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for r in range(3, ws.max_row + 1):
            cell_value = ws.cell(row=r, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


    # 將總計欄位設為粗體
    if '期間總計' in df.columns:
        total_col_idx = df.columns.get_loc('期間總計') + 1
        for row in range(4, ws.max_row + 1):
            ws.cell(row=row, column=total_col_idx).font = Font(bold=True)
            ws.cell(row=row, column=total_col_idx).number_format = '#,##0'

    wb.save(output)
    output.seek(0)
    return output.getvalue()